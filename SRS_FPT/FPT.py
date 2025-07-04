import os
import time
import xmltodict
import pandas as pd
from urllib.parse import urlparse
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook

# Cấu hình trình duyệt
def setup_driver(download_dir):
    os.makedirs(download_dir, exist_ok=True)
    options = Options()
    options.add_experimental_option("prefs", {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True,
        "safebrowsing.enabled": True
    })
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    driver = webdriver.Chrome(options=options)
    return driver, WebDriverWait(driver, 20)

# Đợi file xml xuất hiện
def wait_file_xml(download_dir, timeout=30):
    start_time = time.time()
    while time.time() - start_time < timeout:
        files = [f for f in os.listdir(download_dir) if f.endswith(".xml")]

        # Lọc file đã tải xong (không còn .crdownload)
        files = [f for f in files if not os.path.exists(os.path.join(download_dir, f + ".crdownload"))]
        if files:
            files = sorted(files, key=lambda x: os.path.getmtime(os.path.join(download_dir, x)), reverse=True)
            return os.path.join(download_dir, files[0])
        time.sleep(1)
    return None

# Xử lý tra cứu và tải file tại fpt
def tra_cuu_va_tai_fpt(driver, wait, ma_so_thue, ma_tra_cuu):
    wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='MST bên bán']"))).send_keys(ma_so_thue)
    driver.find_element(By.XPATH, "//input[@placeholder='Mã tra cứu hóa đơn']").send_keys(ma_tra_cuu)
    driver.find_element(By.XPATH, "//button[contains(@class, 'webix_button') and contains(text(), 'Tra cứu')]").click()
    time.sleep(2)
    driver.implicitly_wait(5)
    driver.find_element(By.XPATH, "//button[span[contains(@class, 'mdi-xml')] and contains(text(), 'Tải XML')]").click()
    time.sleep(2)

# Xử lý tra cứu và tải file tại meinvoice
def tra_cuu_va_tai_meinvoice(driver, wait, ma_tra_cuu):
    wait.until(EC.presence_of_element_located((By.NAME, "txtCode"))).send_keys(ma_tra_cuu)
    driver.find_element(By.ID, "btnSearchInvoice").click()
    time.sleep(2)
    driver.find_element(By.CLASS_NAME, "download").click()
    time.sleep(2)
    driver.find_element(By.CLASS_NAME, "txt-download-xml").click()

# Xử lý tra cứu và tải file tại ehoadon
def tra_cuu_va_tai_ehoadon(driver, wait, ma_tra_cuu):
    wait.until(EC.presence_of_element_located((By.ID, "txtInvoiceCode"))).send_keys(ma_tra_cuu)
    driver.find_element(By.CLASS_NAME, "btnSearch").click()
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, "frameViewInvoice")))
    time.sleep(2)
    driver.find_element(By.ID, "btnDownload").click()
    time.sleep(2)
    driver.execute_script("document.querySelector('#divDownloads .dropdown-menu').style.display='block';")
    driver.find_element(By.ID, "LinkDownXML").click()

# Tra cứu và tải XML
def tra_cuu_va_tai_xml(driver, wait, ma_so_thue, ma_tra_cuu, url, download_dir):
    driver.get(url)
    domain = urlparse(url).netloc

    try:
        if "fpt.com.vn" in domain:
            tra_cuu_va_tai_fpt(driver, wait, ma_so_thue, ma_tra_cuu)
        elif "meinvoice.vn" in domain:
            tra_cuu_va_tai_meinvoice(driver, wait, ma_tra_cuu)
        elif "ehoadon.vn" in domain:
            tra_cuu_va_tai_ehoadon(driver, wait, ma_tra_cuu)
        else:
            print(f"Không hỗ trợ URL: {url}")
            return None

        return wait_file_xml(download_dir, timeout=60)

    except Exception as e:
        print(f"Lỗi khi xử lý {url}: {e}")
        return None

# Đọc dữ liệu từ XML
def read_xml_info(xml_path):
    try:
        with open(xml_path, 'r', encoding='utf-8') as f:
            xml_data = f.read()
        data = xmltodict.parse(xml_data)

        invoice_root = None
        candidates = [
            ["HDon", "DLHDon"],
            ["DLHDon"],
            ["TDiep", "DLieu", "HDon", "DLHDon"]
        ]

        for path in candidates:
            current = data
            for key in path:
                if isinstance(current, dict) and key in current:
                    current = current[key]
                else:
                    current = None
                    break
            if current:
                invoice_root = current
                break

        if not isinstance(invoice_root, dict):
            return None

        # Truy cập từng trường theo từng cấp .get()
        ndh = invoice_root.get("NDHDon", {})
        nban = ndh.get("NBan", {})
        nmua = ndh.get("NMua", {})
        ttchung = invoice_root.get("TTChung", {})

        # Xử lý đặc biệt cho STKNHang
        stk_ban_hang = nban.get("STKNHang")
        if not stk_ban_hang:
            ttkhac = nban.get("TTKhac", {}).get("TTin")
            if isinstance(ttkhac, dict):
                ttkhac = [ttkhac]
            for item in ttkhac or []:
                if item.get("TTruong") == "SellerBankAccount":
                    stk_ban_hang = item.get("DLieu")
                    break

        return {
            'Số hóa đơn': ttchung.get("SHDon"),
            'Đơn vị bán hàng': nban.get("Ten"),
            'Mã số thuế bán': nban.get("MST"),
            'Địa chỉ bán': nban.get("DChi"),
            'Số tài khoản bán': stk_ban_hang,
            'Họ tên người mua hàng': nmua.get("Ten"),
            'Địa chỉ mua': nmua.get("DChi"),
            'Mã số thuế mua': nmua.get("MST"),
        }

    except Exception as e:
        print(f"Lỗi đọc XML: {e}")
        return None

# Ghi dữ liệu ra Excel
def write_excel(filepath, data):
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "STT", "Mã số thuế", "Mã tra cứu", "URL",
            "Số hóa đơn", "Đơn vị bán hàng",
            "Mã số thuế bán", "Địa chỉ bán",
            "Số tài khoản bán", "Họ tên người mua hàng",
            "Địa chỉ mua", "Mã số thuế mua", "Ghi chú"
        ])
    else:
        wb = load_workbook(filepath)
        ws = wb.active
    for row in data:
        ws.append(row)
    wb.save(filepath)

# Hàm chính
def main(): 
    input_file = "input.xlsx"
    output_file = "output.xlsx"
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_download_dir = os.path.join(script_dir, "down_invoices")
    os.makedirs(base_download_dir, exist_ok=True)

    df = pd.read_excel(input_file, dtype=str)
    result_rows = []

    for i, row in df.iterrows():
        xml_file = None

        ma_so_thue = str(row.get("Mã số thuế", "")).strip()
        ma_tra_cuu = str(row.get("Mã tra cứu", "")).strip()
        url = str(row.get("URL", "")).strip()
        print(f"\n>> Đang xử lý: {ma_tra_cuu} | Trang: {url}")

        #Phân loại theo domain và mã tra cứu để tạo folder tải xuống
        parsed_url = urlparse(url)
        domain = parsed_url.netloc.lower()

        if "fpt" in domain:
            site_name = "fpt"
        elif "meinvoice" in domain:
            site_name = "meinvoice"
        elif "ehoadon" in domain or "van.ehoadon" in domain:
            site_name = "ehoadon"
        else:
            site_name = "other"

        safe_matracu = "".join(c for c in ma_tra_cuu if c.isalnum() or c in "-_").strip()
        download_dir = os.path.join(base_download_dir, site_name, safe_matracu)
        os.makedirs(download_dir, exist_ok=True)

        #Mở Chrome riêng cho từng dòng url
        driver, wait = setup_driver(download_dir)
        try:
            xml_file = tra_cuu_va_tai_xml(driver, wait, ma_so_thue, ma_tra_cuu, url, download_dir)
        except Exception as e:
            print(f"!! Lỗi khi xử lý dòng {i + 1}: {e}")
        finally:
            driver.quit()
            time.sleep(1)  # nghỉ sau mỗi lần xử lý

        # === Xử lý kết quả ===
        if xml_file:
            info = read_xml_info(xml_file)
            if info:
                result_rows.append([
                    i + 1, ma_so_thue, ma_tra_cuu, url,
                    info.get("Số hóa đơn", ""), info.get("Đơn vị bán hàng", ""),
                    info.get("Mã số thuế bán", ""), info.get("Địa chỉ bán", ""),
                    info.get("Số tài khoản bán", ""), info.get("Họ tên người mua hàng", ""),
                    info.get("Địa chỉ mua", ""), info.get("Mã số thuế mua", "")
                ])
            else:
                result_rows.append([i + 1, ma_so_thue, ma_tra_cuu, url] + [""] * 8 + ["Không đọc được XML"])
        else:
            result_rows.append([i + 1, ma_so_thue, ma_tra_cuu, url] + [""] * 8 + ["Tải XML thất bại"])

    write_excel(output_file, result_rows)
    print(f"\n==> Đã lưu kết quả vào: {output_file}")


if __name__ == "__main__":
    main()
